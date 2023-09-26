from openpyxl import load_workbook
import boto3
import os

from datetime import datetime

s3_client = boto3.client('s3')

def get_excel_data(sheet, header: list) -> list:
    excel_sheet = list()
    # Append header
    excel_sheet.append(header)
    # Iterate through rows and columns
    for row in sheet.iter_rows():
        row_data = []  # Store each row in a list
        for cell in row:
            row_data.append(cell.value) # Append cell value to the row list
        excel_sheet.append(row_data)
    return excel_sheet
    
def get_excel_header(key:str) -> list:
    if 'jobs' in key:
        return ['id', 'job']
    elif 'departments' in key:
        return ['id', 'department']
    elif 'employees' in key:
        return ['id', 'name', 'datetime', 'department_id', 'job_id']
    else:
        return []
        

def lambda_handler(event, context):
    # Getting Bucket name and key
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']
    local_file_path = f'/tmp/{key.split("/")[-1]}'
    
    excel_header = get_excel_header(key)
    
    # getting content
    s3_client.download_file(bucket, key, local_file_path)
    
    # Read the .xlsx file
    workbook = load_workbook(filename=local_file_path)
    sheet = workbook.active
    
    # Retrieve excel data
    excel_sheet = get_excel_data(sheet, excel_header)
    
    # Clean up the downloaded file
    os.remove(local_file_path)

    # Construct the CSV content as a string
    csv_content = '\n'.join([','.join(map(str, row)) for row in excel_sheet])

    # Specify the S3 bucket and key where you want to save the modified JSON
    new_bucket = f"{bucket}-transformed"
    new_key = key.replace("xlsx", "csv")

    # Upload the modified JSON to S3
    s3_client.put_object(Bucket=new_bucket, Key=new_key, Body=csv_content)

    return {
        'statusCode': 200,
        'body': "CSV file saved and uploaded to S3 successfully."
    }